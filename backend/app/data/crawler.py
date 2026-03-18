from dataclasses import dataclass, field
import os
import sqlite3
import logging
from datetime import datetime

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class HskRecord:
    code: str
    name_kr: str
    name_en: str
    level: int
    parent_code: str | None
    description: str
    # 계층 품목명 (신성질별 엑셀용)
    name_2: str = ""   # 세번2단위품명 (류)
    name_4: str = ""   # 세번4단위품명 (호)
    name_6: str = ""   # 세번6단위품명 (소호)
    name_10: str = ""  # 세번10단위품명 (HSK)
    # 신성질별 분류
    category_large: str = ""   # 대분류명
    category_medium: str = ""  # 중분류명
    category_small: str = ""   # 소분류명
    category_detail: str = ""  # 세분류명
    # 임베딩용 결합 텍스트
    full_name: str = ""


class HskCrawler:
    """관세청 HS부호 엑셀 파일에서 HSK 코드를 로드하여 SQLite에 저장"""

    @staticmethod
    def format_code(code: str) -> str:
        """10자리 숫자 코드를 표시 형식으로 변환"""
        code = code.strip()
        if len(code) <= 2:
            return code
        if len(code) == 4:
            return f"{code[:2]}.{code[2:]}"
        if len(code) == 6:
            return f"{code[:4]}.{code[4:]}"
        if len(code) == 8:
            return f"{code[:4]}.{code[4:6]}-{code[6:]}"
        if len(code) == 10:
            return f"{code[:4]}.{code[4:6]}-{code[6:]}"
        return code

    @staticmethod
    def determine_level(code: str) -> int:
        """코드 길이로 계층 수준 결정"""
        length_to_level = {2: 1, 4: 2, 6: 3, 8: 4, 10: 5}
        return length_to_level.get(len(code), 0)

    @staticmethod
    def determine_parent(code: str) -> str | None:
        """부모 코드 결정"""
        parent_lengths = {10: 8, 8: 6, 6: 4, 4: 2}
        parent_len = parent_lengths.get(len(code))
        if parent_len is None:
            return None
        return code[:parent_len]

    @staticmethod
    def _build_full_name(name_2: str, name_4: str, name_6: str, name_10: str,
                         cat_large: str, cat_medium: str, cat_small: str, cat_detail: str) -> str:
        """계층 품목명 + 성질 분류를 결합하여 임베딩용 텍스트 생성"""
        # 계층 품목명 결합: "제85류 전기기기 > 축전지 > 리튬이온 축전지 > 반도체 제조용"
        hierarchy_parts = []
        for name in [name_2, name_4, name_6, name_10]:
            name = name.strip()
            if name and name not in hierarchy_parts:
                hierarchy_parts.append(name)
        hierarchy = " > ".join(hierarchy_parts)

        # 성질 분류 결합: "[자본재 > 전기·전자기기 > 반도체]"
        cat_parts = []
        for cat in [cat_large, cat_medium, cat_small, cat_detail]:
            cat = cat.strip().lstrip("0123456789.)- ").strip()
            if cat and cat not in cat_parts:
                cat_parts.append(cat)
        category = " > ".join(cat_parts)

        if category:
            return f"{hierarchy} [{category}]"
        return hierarchy

    def load_from_excel(self, excel_path: str) -> list[HskRecord]:
        """관세청 HSK별 신성질별 엑셀 파일에서 HSK 코드를 로드.

        엑셀 컬럼 구조 (관세청_HSK별 신성질별_성질별 분류_YYYYMMDD.xlsx):
          B(열1): HS 10자리 부호
          C(열2): 세번2단위품명 (류)
          D(열3): 세번4단위품명 (호)
          E(열4): 세번6단위품명 (소호)
          F(열5): 세번10단위품명 (HSK)
          H(열7): 신성질별 대분류명
          J(열9): 신성질별 중분류명
          L(열11): 신성질별 소분류명
          N(열13): 신성질별 세분류명

        기존 HS부호 엑셀(관세청_HS부호_YYYYMMDD.xlsx)도 호환:
          A(열0): HS부호, D(열3): 한글품목명, E(열4): 영문품목명
        """
        logger.info(f"엑셀 파일 로드 시작: {excel_path}")
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        ws = wb.active

        # 헤더로 파일 형식 감지
        header = [str(cell) if cell else "" for cell in next(ws.iter_rows(max_row=1, values_only=True))]
        is_new_format = "세번2단위품명" in header[2] if len(header) > 2 else False

        records: list[HskRecord] = []
        parent_names: dict[str, str] = {}  # code -> name_kr (상위 코드용)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if is_new_format:
                code = str(row[1]).strip() if row[1] else ""
            else:
                code = str(row[0]).strip() if row[0] else ""

            if not code or len(code) != 10 or not code.isdigit():
                continue

            if is_new_format:
                name_2 = str(row[2]).strip() if row[2] else ""
                name_4 = str(row[3]).strip() if row[3] else ""
                name_6 = str(row[4]).strip() if row[4] else ""
                name_10 = str(row[5]).strip() if row[5] else ""
                cat_large = str(row[7]).strip() if row[7] else ""
                cat_medium = str(row[9]).strip() if row[9] else ""
                cat_small = str(row[11]).strip() if row[11] else ""
                cat_detail = str(row[13]).strip() if row[13] else ""

                full_name = self._build_full_name(
                    name_2, name_4, name_6, name_10,
                    cat_large, cat_medium, cat_small, cat_detail,
                )
                name_kr = name_10 if name_10 else name_6
                name_en = ""
                description = ""
            else:
                name_kr = str(row[3]).strip() if row[3] else ""
                name_en = str(row[4]).strip() if row[4] else ""
                description = str(row[5]).strip() if row[5] else ""
                name_2 = name_4 = name_6 = name_10 = ""
                cat_large = cat_medium = cat_small = cat_detail = ""
                full_name = f"{name_kr} ({name_en})" if name_en else name_kr

            record = HskRecord(
                code=code,
                name_kr=name_kr,
                name_en=name_en,
                level=5,
                parent_code=code[:8],
                description=description,
                name_2=name_2,
                name_4=name_4,
                name_6=name_6,
                name_10=name_10,
                category_large=cat_large,
                category_medium=cat_medium,
                category_small=cat_small,
                category_detail=cat_detail,
                full_name=full_name,
            )
            records.append(record)

            # 상위 코드 품명 수집
            if is_new_format:
                if code[:2] and name_2:
                    parent_names[code[:2]] = name_2
                if code[:4] and name_4:
                    parent_names[code[:4]] = name_4
                if code[:6] and name_6:
                    parent_names[code[:6]] = name_6

        wb.close()

        # 상위 계층 코드 생성
        existing_codes = {r.code for r in records}
        parent_records = []
        parent_codes_seen = set()
        for r in records:
            for length in [2, 4, 6, 8]:
                parent_codes_seen.add(r.code[:length])

        for pcode in sorted(parent_codes_seen):
            if pcode not in existing_codes:
                pname = parent_names.get(pcode, f"[{self.format_code(pcode)}]")
                parent_records.append(HskRecord(
                    code=pcode,
                    name_kr=pname,
                    name_en="",
                    level=self.determine_level(pcode),
                    parent_code=self.determine_parent(pcode),
                    description="",
                    full_name=pname,
                ))

        all_records = parent_records + records
        logger.info(f"엑셀 로드 완료: HSK {len(records)}건 + 상위코드 {len(parent_records)}건 = 총 {len(all_records)}건")
        return all_records

    def save_to_sqlite(self, records: list[HskRecord], db_path: str, source_file: str = "") -> None:
        """수집한 레코드를 SQLite에 저장하고 데이터 소스 이력 기록"""
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        cursor.execute("DROP TABLE IF EXISTS hsk_codes")
        cursor.execute("""
            CREATE TABLE hsk_codes (
                code TEXT PRIMARY KEY,
                name_kr TEXT NOT NULL,
                name_en TEXT,
                level INTEGER NOT NULL,
                parent_code TEXT,
                description TEXT,
                name_2 TEXT,
                name_4 TEXT,
                name_6 TEXT,
                name_10 TEXT,
                category_large TEXT,
                category_medium TEXT,
                category_small TEXT,
                category_detail TEXT,
                full_name TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS data_sources (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_name TEXT NOT NULL,
                loaded_at TEXT NOT NULL,
                record_count INTEGER NOT NULL
            )
        """)

        cursor.executemany(
            """INSERT INTO hsk_codes
               (code, name_kr, name_en, level, parent_code, description,
                name_2, name_4, name_6, name_10,
                category_large, category_medium, category_small, category_detail,
                full_name)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [(r.code, r.name_kr, r.name_en, r.level, r.parent_code, r.description,
              r.name_2, r.name_4, r.name_6, r.name_10,
              r.category_large, r.category_medium, r.category_small, r.category_detail,
              r.full_name) for r in records],
        )

        if source_file:
            cursor.execute(
                "INSERT INTO data_sources (file_name, loaded_at, record_count) VALUES (?, ?, ?)",
                (os.path.basename(source_file), datetime.now().isoformat(), len(records)),
            )

        conn.commit()
        conn.close()
        logger.info(f"SQLite 저장 완료: {len(records)}건 → {db_path}")
