import json
import logging
from openpyxl import Workbook, load_workbook
from app.data.batch_db import BatchDB

logger = logging.getLogger(__name__)

MAX_ROWS = 500
MIN_DESC_LENGTH = 10


class BatchService:
    def __init__(self, batch_db: BatchDB):
        self.db = batch_db

    def parse_excel(self, file_path: str) -> list[dict]:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        items = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 2:
                continue
            task_name = str(row[0]).strip() if row[0] else None
            description = str(row[1]).strip() if row[1] else ""
            if len(description) < MIN_DESC_LENGTH:
                continue
            items.append({"row_index": row_idx, "task_name": task_name, "description": description})
        wb.close()
        if len(items) > MAX_ROWS:
            raise ValueError(f"최대 {MAX_ROWS}건까지 지원합니다. (입력: {len(items)}건)")
        return items

    def create_template(self, output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "입력"
        ws.append(["과제명", "기술설명"])
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 80
        wb.save(output_path)

    def create_job(self, file_path: str, file_name: str, top_n: int,
                   confidence_threshold: float | None) -> str:
        items = self.parse_excel(file_path)
        if not items:
            raise ValueError("유효한 기술 설명이 없습니다.")
        job_id = self.db.create_job(file_name, len(items), top_n, confidence_threshold)
        self.db.create_items(job_id, items)
        return job_id

    def generate_result_excel(self, job_id: str, output_path: str):
        job = self.db.get_job(job_id)
        items = self.db.get_items(job_id)
        top_n = job["top_n"]
        confidence_threshold = job["confidence_threshold"]

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "요약"

        parsed_items = []
        max_codes = 0
        for item in items:
            result_data = json.loads(item["result_json"]) if item["result_json"] else None
            codes = []
            if result_data and item["status"] == "completed":
                results = result_data.get("results", [])
                if confidence_threshold is not None:
                    results = [r for r in results if r.get("confidence", 0) >= confidence_threshold]
                codes = [r["hsk_code"] for r in results]
            keywords = ", ".join(result_data.get("keywords_extracted", [])) if result_data else ""
            parsed_items.append({
                "item": item, "codes": codes, "keywords": keywords,
                "results": result_data.get("results", []) if result_data else [],
            })
            max_codes = max(max_codes, len(codes))

        if confidence_threshold is None:
            num_code_cols = top_n
        else:
            num_code_cols = min(max_codes, 30)

        headers = ["과제명", "기술설명", "상태", "추출 키워드"]
        for i in range(1, num_code_cols + 1):
            headers.append(f"HSK코드_{i}")
        ws_summary.append(headers)

        for p in parsed_items:
            row = [
                p["item"]["task_name"] or "",
                p["item"]["description"],
                "성공" if p["item"]["status"] == "completed" else "실패",
                p["keywords"],
            ]
            for i in range(num_code_cols):
                row.append(p["codes"][i] if i < len(p["codes"]) else "")
            ws_summary.append(row)

        ws_detail = wb.create_sheet("상세")
        ws_detail.append(["과제명", "기술설명", "순위", "HSK코드", "품목명(한)", "품목명(영)", "신뢰도(%)", "선정 사유"])

        for p in parsed_items:
            item = p["item"]
            if item["status"] == "completed" and p["results"]:
                results = p["results"]
                if confidence_threshold is not None:
                    results = [r for r in results if r.get("confidence", 0) >= confidence_threshold]
                for r in results:
                    ws_detail.append([
                        item["task_name"] or "",
                        item["description"],
                        r.get("rank", ""),
                        r.get("hsk_code", ""),
                        r.get("name_kr", ""),
                        r.get("name_en", ""),
                        round(r.get("confidence", 0) * 100, 1),
                        r.get("reason", ""),
                    ])
            elif item["status"] == "failed":
                ws_detail.append([
                    item["task_name"] or "",
                    item["description"],
                    "에러", "", "", "", "",
                    item.get("error_message", "알 수 없는 오류"),
                ])

        wb.save(output_path)
