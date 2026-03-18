import json
import pytest
from openpyxl import Workbook
from app.services.batch_service import BatchService
from app.data.batch_db import BatchDB


@pytest.fixture
def db(tmp_path):
    return BatchDB(str(tmp_path / "test.db"))


@pytest.fixture
def service(db):
    return BatchService(db)


def _make_excel(tmp_path, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(["과제명", "기술설명"])
    for row in rows:
        ws.append(row)
    path = str(tmp_path / "test.xlsx")
    wb.save(path)
    return path


def test_parse_excel(tmp_path, service):
    path = _make_excel(tmp_path, [
        ["과제A", "리튬이온 배터리 양극재 제조 기술에 대한 설명입니다"],
        ["과제B", "수소연료전지 막전극접합체 제조를 위한 촉매 기술"],
        [None, ""],
        [None, "짧은"],
    ])
    items = service.parse_excel(path)
    assert len(items) == 2
    assert items[0]["task_name"] == "과제A"
    assert items[0]["row_index"] == 2


def test_parse_excel_max_rows(tmp_path, service):
    rows = [[f"과제{i}", f"기술 설명 번호 {i} 입니다. 이것은 충분히 긴 설명입니다."] for i in range(501)]
    path = _make_excel(tmp_path, rows)
    with pytest.raises(ValueError, match="500건"):
        service.parse_excel(path)


def test_create_template(tmp_path, service):
    path = str(tmp_path / "template.xlsx")
    service.create_template(path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    assert ws.cell(1, 1).value == "과제명"
    assert ws.cell(1, 2).value == "기술설명"


def test_create_job_from_excel(tmp_path, db, service):
    path = _make_excel(tmp_path, [
        ["과제A", "리튬이온 배터리 양극재 제조 기술에 대한 설명입니다"],
    ])
    job_id = service.create_job(path, "test.xlsx", top_n=5, confidence_threshold=None, model="chatgpt-5.4-mini")
    job = db.get_job(job_id)
    assert job["total_items"] == 1
    items = db.get_items(job_id)
    assert len(items) == 1


def test_generate_result_excel_topn_mode(tmp_path, db, service):
    job_id = db.create_job("test.xlsx", 1, 5, None, "chatgpt-5.4-mini")
    db.create_items(job_id, [{"row_index": 2, "task_name": "과제A", "description": "기술 설명"}])
    items = db.get_items(job_id)
    result = {
        "results": [
            {"rank": 1, "hsk_code": "8507.60-1000", "name_kr": "리튬이온 축전지", "name_en": "Li-ion", "confidence": 0.95, "reason": "사유"},
            {"rank": 2, "hsk_code": "8507.50-0000", "name_kr": "니켈 축전지", "name_en": None, "confidence": 0.7, "reason": "사유2"},
        ],
        "keywords_extracted": ["양극재", "배터리"],
    }
    db.update_item_status(items[0]["item_id"], "completed", result_json=json.dumps(result, ensure_ascii=False))
    db.refresh_job_progress(job_id)

    output_path = str(tmp_path / "result.xlsx")
    service.generate_result_excel(job_id, output_path)

    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    assert "요약" in wb.sheetnames
    assert "상세" in wb.sheetnames
    summary = wb["요약"]
    assert summary.cell(1, 5).value == "HSK코드_1"
    assert summary.cell(2, 5).value == "8507.60-1000"
    detail = wb["상세"]
    assert detail.cell(2, 4).value == "8507.60-1000"
    assert detail.cell(3, 4).value == "8507.50-0000"
